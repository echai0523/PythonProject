# -*- coding: utf-8 -*-
"""
Author: Ethan Chai
Date: 2022/10/1 10:03
input:
    - txt所在文件夹路径
output:
    - excel包
Short Description:
    - 读取txt转为excel
Change History:
"""
import os
import glob
from openpyxl import Workbook


# 月映射当月天数
month_day_map = dict()
day_list = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
for i, j in zip([f'{k + 1}月' for k in range(12)], day_list):
    month_day_map[i] = j


def function():
    # 获取全部txt文件
    files = glob.glob(os.path.join(folder_path, '*.txt'))

    for file in files:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        # 先写好excel模版
        index = 2  # 第二行开始
        for month, days in month_day_map.items():
            # A2 A33 ··· 写月
            ws[f'A{index}'] = month
            # B2 B3 ··· 写日
            for day in range(days):
                ws[f'B{index + day}'] = day + 1
            # index做好递增
            index += days

        # 后续-年、值-均从C开始写
        letter = "C"
        letter_pre = ""  # 超过Z需要加prefix

        with open(file, mode='r', encoding='utf-8') as readfile:
            # 读取所有行，每行去掉两头的空格、换行符，切片去掉txt头部从第二行取
            lines = [line.strip() for line in readfile.readlines()][1:]

        # lines分组
        lines_of_groups = [lines[ind:ind+14] for ind in range(0, len(lines), 14)]
        for idx, group in enumerate(lines_of_groups):
            # 分析letter、letter_pre取值
            if ord(letter) > ord("Z"):
                letter = "A"
                if letter_pre == "":
                    letter_pre = "A"
                else:
                    # 如果letter_pre已经是大写字母，则直接递增
                    letter_pre = chr(ord(letter) + 1)

            num_year, *lines_value = group
            # 年份
            year = num_year.split()[1]
            ws[f'{letter_pre}{letter}1'] = year  # 年份只在第1行写
            # 值
            values = ' '.join(lines_value).split()
            for value_idx, value in enumerate(values):
                ws[f'{letter_pre}{letter}{2+value_idx}'] = int(value)  # 从第2行开始写

            # letter 递增
            letter = chr(ord(letter) + 1)

        # 保存excel
        file_name = os.path.basename(file).replace('.txt', '.xlsx')
        wb.save(os.path.join(excel_save_path, file_name))


# txt包路径
folder_path = '../../EthanFileData/File txt/伊犁河流域mci'
# 输出的Excel包路径
excel_save_path = '../EthanFileData/File txt/伊犁河流域mci_excel'
os.makedirs(excel_save_path, exist_ok=True)

function()
