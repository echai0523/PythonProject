# -*- coding: UTF-8 -*-

"""
Author: Ethan Chai
Date: 2022/8/17 20:14

input:
    - 学生信息表.xlsx
output:
    - 统计表.xlsx
Short Description: 
    - 知识点：
        - openpyxl
        - ws.merge_cells("A1:A3")
    - 需求：
        - 抽取'学生信息表'的某些列，放到'统计表'指定列，并且前三行合并单元格，数据从第四行开始写入。
Change History:

"""
import pandas as pd
from openpyxl import Workbook


def main():
    info_df = pd.read_excel(info_path, sheet_name="Sheet1")

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    # 单元格合并
    for a_h, value in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'],
                          ['学号', '姓名', '性别', '此列空着', '此列空着', '年级', '此列空着', '籍贯']):
        # 先写入值
        ws[f"{a_h}1"] = value
        # 再合并
        ws.merge_cells(f"{a_h}1:{a_h}3")

    for index, row in info_df.iterrows():
        ws[f"A{index + 4}"] = row["学号"]
        ws[f"B{index + 4}"] = row["姓名"]
        ws[f"C{index + 4}"] = row["性别"]
        ws[f"F{index + 4}"] = row["年级"]
        ws[f"H{index + 4}"] = row["籍贯"]

    wb.save(count_path)


if __name__ == '__main__':
    info_path = "../../EthanFileData/xlsx/学生信息表.xlsx"
    count_path = "../../EthanFileData/xlsx/统计表.xlsx"

    main()
