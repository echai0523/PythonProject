# -*- coding: UTF-8 -*-

"""
Author: Ethan Chai
Date: 2022/8/23 13:09

input: 
output: 
Short Description: 

Change History:

"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Color, Font, Alignment, PatternFill
from peutils.transform.v1.audio_seg.parser import AudioSegParse, AudioSegDataConfig
import sys
from peutils.wooeyutil import WooeyBaseZipHandlerFile
from pprint import pprint


def main(csv_file_path):
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    # 合并单元格
    ws.merge_cells("A1:F1")
    top_label = ws['A1']
    top_label.value = "MOS评测标注"
    # font字体，fill填充(单元格背景)，alignment居中
    top_label.font = Font(name=u'黑体')
    top_label.fill = PatternFill("solid", fgColor="B4C6E7")
    top_label.alignment = Alignment(horizontal="center", vertical="center")
    # align = Alignment(horizontal='center', vertical='center')
    ws.append(['标注编号', '模型编号', '音频编号', '音质', '自然度', '情感表现力'])
    # column_name = ws['A2:F2']字体转为黑体
    for s in [chr(A) for A in range(65, 71)]:
        ws[f'{s}2'].font = Font(name=u'黑体')
    df = pd.read_csv(csv_file_path)

    for _, row in df.iterrows():
        model_num, audio_name = row["audio_url"].split('/')[-2:]
        audio_num = audio_name[:-len(".wav")]
        ad = AudioSegParse(
            url=row["annotation"],
            config=AudioSegDataConfig(segment_mode="continuous")
        )
        assert ad.frame_length == 1, "音频帧数不为1"
        attr = ad.frames_lst[0].frame_attr
        ws.append([row["Annotation Worker Name"], model_num, audio_num, attr["音质"], attr["自然度"], attr["情感表现力"]])

    # 整个excel进行居中对齐
    max_rows = ws.max_row  # 获取最大行
    max_columns = ws.max_column  # 获取最大列
    for i in range(1, max_rows + 1):
        for j in range(1, max_columns + 1):
            ws.cell(i, j).alignment = Alignment(horizontal="center", vertical="center")
    wb.save("MOS评测标注.xlsx")

    pass


if __name__ == '__main__':
    # main(r"C:\Users\echai\Downloads\process\1518 音频\result_report_A6014-Li133292_2022-08-23T03_21_37.371453ZUTC.csv")
    h = WooeyBaseZipHandlerFile('MOS评测标注', '.csv', process_func=main)
    sys.exit(h.main())
